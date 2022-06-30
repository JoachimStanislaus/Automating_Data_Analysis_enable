from openpyxl import Workbook, load_workbook
import pandas as pd

wb = load_workbook(filename='Sample Data.xlsx')
ws = wb.active

new_wb = Workbook()
new_ws = new_wb.active

data_dict = {}
data_list = []
counter = 1

def read_row(worksheet, row): # Read row cell by cell and return cell values in a list
    return [cell.value for cell in worksheet[row]]

def remove_empty_columns(row_list): # Checks if Cell is None and removes it if it is. Returns a list without None cells.
    emptylist=[]
    for x in row_list: 
        if x is not None:
            emptylist.append(x)
    return(emptylist)

def save_to_list(row_list): # saving data to a list
    emptylist=[]
    for x in row_list:
        emptylist.append(x)
    return(emptylist)

def fill_empty_cell(): # Fill the empty cells that are used in final analysis 
    for x in range(1,len(data_dict)+1):
        for i in range(0,len(data_dict[1])):
            if data_dict[x][i] is None:
                if x <len(data_dict) and x>1: # Check if it's not the first or last row
                    if (isinstance(data_dict[x-1][i], int)) == True or (isinstance(data_dict[x+1][i], int)) == True: # Checking if the upper or lower row cell is filled with a int
                        data_dict[x][i] = 0 # Filling the cell with 0 if the upper or lower cell is filled with a int

                    elif (isinstance(data_dict[x-1][i], str)) == True or (isinstance(data_dict[x+1][i], str)) == True: # Checking if the upper or lower row cell is filled with a str
                        data_dict[x][i] = 'Nil' # Filling the cell with Nil if the upper or lower cell is filled with a str

                elif x==len(data_dict): # Check if it's the last row
                    if (isinstance(data_dict[x-1][i], int)) == True or (isinstance(data_dict[x-2][i], int)) == True: # Checking if the upper or lower row cell is filled with a int
                        data_dict[x][i] = 0 # Filling the cell with 0 if the upper or lower cell is filled with a int

                    elif (isinstance(data_dict[x-1][i], str)) == True or (isinstance(data_dict[x-2][i], str)) == True: # Checking if the upper or lower row cell is filled with a str
                        data_dict[x][i] = 'Nil' # Filling the cell with Nil if the upper or lower cell is filled with a str

                elif x==1 : # Check if it's the first row 
                    if (isinstance(data_dict[x+1][i], int)) == True or (isinstance(data_dict[x+2][i], int)) == True: # Checking if the upper or lower row cell is filled with a int
                        data_dict[x][i] = 0 # Filling the cell with 0 if the upper or lower cell is filled with a int

                    elif (isinstance(data_dict[x+1][i], str)) == True or (isinstance(data_dict[x+2][i], str)) == True: # Checking if the upper or lower row cell is filled with a str
                        data_dict[x][i] = 'Nil' # Filling the cell with Nil if the upper or lower cell is filled with a str

def populate_sheet(worksheet, row,row_data):
    for i in range(1,len(row_data)+1): #loop through all the columns and populate the sheet
            cellref=worksheet.cell(row=row, column=i) #declaring the cell to be given value
            cellref.value=row_data[i-1] #add value to cell

for x in range(1,50): # Cycle through all the rows
    row_data = read_row(ws,x) # Get row_data row by row
    if  (isinstance(row_data[0], int)) == True : # check if the first column is a number if it is means the data is useful data
        print(row_data)
        data_dict[counter] = row_data
        counter+=1
fill_empty_cell()

for x in range(1,len(data_dict)+1): # Loop through the every single row and remove empty columns
    data_dict[x] = remove_empty_columns(data_dict[x])
    populate_sheet(new_ws,x,data_dict[x]) # create new sheet and populate sheet with data row by row

new_wb.save('Clean_Data.xlsx') # Save data to New Excel Workbook

# oneliner
pd.read_excel('Clean_Data.xlsx').to_csv('Clean_Dataa.csv', index=False) # Uses Pandas to convert new workbook to csv file.
